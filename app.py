"""
Energybae Solar Load Calculator - Streamlit Web App
Reads MSEDCL electricity bills and generates Solar Load Calculator Excel report.
"""

import streamlit as st
import base64
import json
import io
import re
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="Energybae Solar Load Calculator",
    page_icon="☀️",
    layout="centered"
)

# ============================================================
# API KEY - Loaded securely from Streamlit Secrets
# ============================================================
try:
    API_KEY = st.secrets["OPENROUTER_API_KEY"]
except Exception:
    st.error("⚠️ API Key not configured. Please set OPENROUTER_API_KEY in Streamlit Secrets.")
    st.stop()


# ============================================================
# BILL EXTRACTION FUNCTION (from your notebook - Cell 11)
# ============================================================
def extract_bill_data(file_bytes, media_type, api_key):
    client = OpenAI(
        api_key=api_key,
        base_url="https://openrouter.ai/api/v1"
    )

    prompt = """You are an expert at reading MSEDCL electricity bills.
Extract all fields and return ONLY valid JSON, no markdown, no explanation.
Keep all string values SHORT (under 50 characters).
Return exactly this structure:
{
  "consumer_name": "...",
  "consumer_number": "...",
  "billing_unit": "...",
  "tariff_rate": "...",
  "meter_number": "...",
  "reading_group": "...",
  "sanctioned_load_kw": 0,
  "security_deposit": 0,
  "bill_month": "...",
  "bill_date": "...",
  "due_date": "...",
  "current_reading": 0,
  "previous_reading": 0,
  "units_consumed": 0,
  "total_bill_amount": 0,
  "monthly_units": {
    "Feb-2025": 0, "Mar-2025": 0, "Apr-2025": 0, "May-2025": 0,
    "Jun-2025": 0, "Jul-2025": 0, "Aug-2025": 0, "Sep-2025": 0,
    "Oct-2025": 0, "Nov-2025": 0, "Dec-2025": 0, "Jan-2026": 0
  }
}"""

    b64 = base64.standard_b64encode(file_bytes).decode()
    image_url = f"data:{media_type};base64,{b64}"

    response = client.chat.completions.create(
        model="nvidia/nemotron-nano-12b-v2-vl:free",
        max_tokens=2000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": image_url}}
            ]
        }],
        extra_headers={
            "HTTP-Referer": "https://energybae.in",
            "X-Title": "Energybae Solar Calculator"
        }
    )

    text = response.choices[0].message.content.strip()

    # Clean markdown fences if present
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    text = text.strip()

    # Try parsing JSON, with fallback for incomplete responses
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        if not text.endswith("}"):
            text = text + '}}'
        try:
            return json.loads(text)
        except Exception:
            # Manual fallback - extract key-value pairs with regex
            numbers = re.findall(r'"(\w[^"]+)":\s*(\d+\.?\d*)', text)
            strings = re.findall(r'"(\w[^"]+)":\s*"([^"]*)"', text)
            result = {}
            for k, v in strings:
                result[k] = v
            for k, v in numbers:
                result[k] = float(v) if '.' in v else int(v)
            if "monthly_units" not in result:
                result["monthly_units"] = {}
            return result


# ============================================================
# EXCEL GENERATOR FUNCTION (from your notebook - Cell 16)
# ============================================================
def create_solar_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load Analysis"

    def cs(ref, val, bold=False, bg=None, fc="000000", align="left", size=10):
        c = ws[ref]
        c.value = val
        c.font = Font(bold=bold, color=fc, size=size, name="Arial")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        t = Side(style="thin", color="BDBDBD")
        c.border = Border(left=t, right=t, top=t, bottom=t)

    def mt(rng, val, bg="1B5E20", fc="FFFFFF", size=12):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = val
        c.font = Font(bold=True, color=fc, size=size, name="Arial")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=bg)

    for col, w in {"A": 30, "B": 22, "C": 30, "D": 22, "E": 15, "F": 15}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 38
    mt("A1:F1", "ENERGYBAE — SOLAR LOAD CALCULATOR REPORT", size=14)
    mt("A2:F2",
       f"Consumer: {data.get('consumer_name', 'N/A')}  |  Bill: {data.get('bill_month', 'N/A')}",
       bg="2E7D32", size=10)

    row = 4
    mt(f"A{row}:F{row}", "SECTION 1 — CONSUMER DETAILS", bg="388E3C", size=11)
    row += 1

    for l1, v1, l2, v2 in [
        ("Consumer Name", data.get("consumer_name", "N/A"),
         "Consumer Number", data.get("consumer_number", "N/A")),
        ("Billing Unit", data.get("billing_unit", "N/A"),
         "Tariff Code", data.get("tariff_rate", "N/A")),
        ("Meter Number", data.get("meter_number", "N/A"),
         "Reading Group", data.get("reading_group", "N/A")),
        ("Sanctioned Load(kW)", data.get("sanctioned_load_kw", "N/A"),
         "Security Deposit", data.get("security_deposit", "N/A")),
        ("Bill Date", data.get("bill_date", "N/A"),
         "Due Date", data.get("due_date", "N/A")),
        ("Bill Month", data.get("bill_month", "N/A"),
         "Total Bill (Rs.)", data.get("total_bill_amount", "N/A")),
    ]:
        ws.row_dimensions[row].height = 22
        cs(f"A{row}", l1, bold=True, bg="C8E6C9")
        cs(f"B{row}", v1, bg="FFFFFF", align="center")
        cs(f"C{row}", l2, bold=True, bg="C8E6C9")
        cs(f"D{row}", v2, bg="FFFFFF", align="center")
        ws.merge_cells(f"E{row}:F{row}")
        row += 1

    row += 1
    mt(f"A{row}:F{row}", "SECTION 2 — METER READING", bg="388E3C", size=11)
    row += 1
    for i, (lbl, val) in enumerate([
        ("Current Reading", data.get("current_reading", "N/A")),
        ("Previous Reading", data.get("previous_reading", "N/A")),
        ("Units Consumed", data.get("units_consumed", "N/A")),
    ]):
        c1 = get_column_letter(i * 2 + 1)
        c2 = get_column_letter(i * 2 + 2)
        cs(f"{c1}{row}", lbl, bold=True, bg="C8E6C9", align="center")
        cs(f"{c2}{row}", val, bg="FFF9C4", align="center", bold=True, size=12)
    row += 1

    row += 1
    mt(f"A{row}:F{row}", "SECTION 3 — 12-MONTH HISTORY", bg="388E3C", size=11)
    row += 1
    for i, h in enumerate(["Month", "Units", "Month", "Units", "Month", "Units"], start=1):
        cs(f"{get_column_letter(i)}{row}", h, bold=True, bg="4CAF50", fc="FFFFFF", align="center")
    row += 1

    months = ["Feb-2025", "Mar-2025", "Apr-2025", "May-2025", "Jun-2025", "Jul-2025",
              "Aug-2025", "Sep-2025", "Oct-2025", "Nov-2025", "Dec-2025", "Jan-2026"]
    monthly = data.get("monthly_units", {})

    for i in range(4):
        ws.row_dimensions[row].height = 20
        for j in range(3):
            idx = i * 3 + j
            m = months[idx]
            v = monthly.get(m, "N/A")
            cs(f"{get_column_letter(j * 2 + 1)}{row}", m, bg="E3F2FD", align="center")
            cs(f"{get_column_letter(j * 2 + 2)}{row}", v, bg="FFFFFF", align="center", bold=(idx == 11))
        row += 1

    row += 1
    mt(f"A{row}:F{row}", "SECTION 4 — SOLAR RECOMMENDATION", bg="E65100", fc="FFFFFF", size=11)
    row += 1

    vals = [monthly.get(m) for m in months if monthly.get(m) not in (None, 0)]
    avg = sum(vals) / len(vals) if vals else 0
    avg_daily = avg / 30
    solar_kw = round(avg_daily / 4.5 * 1.25, 2)
    panels = max(1, round((solar_kw * 1000) / 450))
    actual_kw = round(panels * 450 / 1000, 2)
    savings = round(avg * 12 * 7.5)
    payback = round(actual_kw * 60000 / savings, 1) if savings else "N/A"
    co2 = round(avg * 12 * 0.82)

    for lbl, val, bg in [
        ("Avg Monthly Units (kWh)", round(avg, 1), "FFF3E0"),
        ("Avg Daily Consumption (kWh/day)", round(avg_daily, 2), "FFF9C4"),
        ("Recommended Solar Capacity (kW)", solar_kw, "C8E6C9"),
        ("Panel Wattage Assumed", "450W Mono PERC", "E3F2FD"),
        ("No. of Panels Required", panels, "FFF3E0"),
        ("Total Installed Capacity (kW)", actual_kw, "C8E6C9"),
        ("System Type", "On-Grid", "E3F2FD"),
        ("Est. Annual Savings (Rs.)", f"Rs.{savings:,}", "C8E6C9"),
        ("Payback Period", f"{payback} yrs", "FFF9C4"),
        ("CO2 Offset/Year (kg)", co2, "E3F2FD"),
    ]:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:C{row}")
        cs(f"A{row}", lbl, bold=True, bg=bg, size=10)
        ws.merge_cells(f"D{row}:F{row}")
        cs(f"D{row}", val, bold=True, bg="FFFFFF", align="center", size=11)
        row += 1

    row += 1
    mt(f"A{row}:F{row}",
       "www.energybae.in  |  energybae.co@gmail.com  |  +91 9112233120",
       bg="1B5E20", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# STREAMLIT UI
# ============================================================
st.title("☀️ Energybae Solar Load Calculator")
st.markdown(
    "**Upload your MSEDCL electricity bill** and get an AI-generated "
    "Solar Load Calculator Excel report instantly."
)
st.divider()

uploaded_file = st.file_uploader(
    "📤 Upload Electricity Bill (JPG, PNG, or PDF)",
    type=["jpg", "jpeg", "png", "pdf"]
)

if uploaded_file is not None:
    file_bytes = uploaded_file.read()
    filename = uploaded_file.name
    ext = filename.split(".")[-1].lower()
    mime_map = {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png"
    }
    media_type = mime_map.get(ext, "image/jpeg")

    st.success(f"✅ File uploaded: **{filename}**")

    # Preview image if it's an image file
    if ext in ["jpg", "jpeg", "png"]:
        st.image(file_bytes, caption="Uploaded Bill Preview", use_container_width=True)

    if st.button("🚀 Generate Solar Report", type="primary", use_container_width=True):
        with st.spinner("🤖 AI is reading your bill... please wait (30-60 seconds)"):
            try:
                data = extract_bill_data(file_bytes, media_type, API_KEY)
            except Exception as e:
                st.error(f"❌ Error extracting data: {e}")
                st.stop()

        st.success("✅ Bill data extracted successfully!")

        # Show extracted data preview
        with st.expander("📋 View Extracted Data", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Consumer Name", data.get("consumer_name", "N/A"))
                st.metric("Bill Month", data.get("bill_month", "N/A"))
                st.metric("Units Consumed", data.get("units_consumed", "N/A"))
            with col2:
                st.metric("Consumer Number", data.get("consumer_number", "N/A"))
                st.metric("Total Bill (Rs.)", data.get("total_bill_amount", "N/A"))
                st.metric("Sanctioned Load (kW)", data.get("sanctioned_load_kw", "N/A"))

            st.markdown("**Monthly Consumption:**")
            monthly = data.get("monthly_units", {})
            if monthly:
                st.json(monthly)

        # Generate Excel
        with st.spinner("📊 Creating Excel report..."):
            excel_bytes = create_solar_excel(data)

        name = str(data.get("consumer_name", "Consumer")).replace(" ", "_")
        month = str(data.get("bill_month", "")).replace(" ", "_")
        out_filename = f"Solar_Report_{name}_{month}.xlsx"

        st.success("🎉 Solar Report Ready! Click below to download.")
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_bytes,
            file_name=out_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()
st.caption(
    "Built for **Energybae** — Empowering People with Renewable Energy  \n"
    "🌐 www.energybae.in | ✉️ energybae.co@gmail.com | 📞 +91 9112233120"
)
